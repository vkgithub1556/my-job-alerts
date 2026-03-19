"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  RESUME-POWERED SALESFORCE JOB ALERT SYSTEM v2.0                           ║
║  • Drop ANY resume (PDF or DOCX) into the /resumes folder                  ║
║  • Auto-reads resume, extracts skills, tailors job search + scoring        ║
║  • Emails multi-sheet Excel every 6 hours with ONLY NEW jobs               ║
║  SHEETS: 🆕 New Jobs | 🌲 Alberta Only | 🏆 Top Matches | 📊 Dashboard     ║
╚══════════════════════════════════════════════════════════════════════════════╝

REPO STRUCTURE:
  salesforce-job-alerts/
  ├── job_alert_runner_v2.py        ← this file (root)
  ├── resumes/
  │   └── my_resume.pdf            ← DROP YOUR RESUME HERE (PDF or DOCX)
  ├── seen_jobs.json                ← auto-created, tracks sent jobs
  └── .github/
      └── workflows/
          └── job_alert.yml

GITHUB SECRETS:
  APIFY_API_TOKEN  |  GMAIL_ADDRESS  |  GMAIL_APP_PASSWORD  |  RECIPIENT_EMAIL
"""

import os, json, re, smtplib, sys, glob
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from apify_client import ApifyClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ════════════════════════════════════════════════════════════════════════
DARK_NAVY = "0D1B2A"; GOLD = "C9A84C"; WHITE = "FFFFFF"
GRAY1 = "F5F5F5";    GRAY2 = "EAEAEA"; AB_GREEN = "E8F5E9"; AB_DARK = "1A5F3A"
TIER_COLORS = {
    "🏆 EXCELLENT": ("1A7340","FFFFFF"), "✅ STRONG":    ("2D6A2D","FFFFFF"),
    "🟡 GOOD":      ("7F6000","FFFFFF"), "🟠 STRETCH":   ("8B3A0F","FFFFFF"),
    "❌ LOW MATCH": ("5A0000","FFFFFF"),
}
PRI = {
    "🏆 EXCELLENT":"🚀 APPLY NOW",  "✅ STRONG":"✅ Apply – Tailor CV",
    "🟡 GOOD":"📝 Worth Applying",  "🟠 STRETCH":"⚠️ Optional",
    "❌ LOW MATCH":"❌ Skip",
}
SEEN_FILE = "seen_jobs.json"

# ════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ════════════════════════════════════════════════════════════════════════
def tb(c):
    s = Side(style="thin", color="CCCCCC")
    c.border = Border(left=s, right=s, top=s, bottom=s)

def banner(ws, row, total_cols, text, bg=DARK_NAVY, fg=GOLD, sz=12, height=28):
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(bold=True, color=fg, name="Arial", size=sz)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height

def col_headers(ws, row, headers, bg="1E3A5F", fg=GOLD, height=34):
    ws.row_dimensions[row].height = height
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(bold=True, color=fg, name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        tb(c)

def data_cell(ws, row, col, val, bg, fg=None, bold=False, center=False, underline=False):
    c = ws.cell(row=row, column=col, value=str(val) if val is not None else "")
    c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(name="Arial", size=9, color=fg or "000000", bold=bold,
                  underline="single" if underline else None)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    tb(c)
    return c

def write_job_row(ws, row, vals, tier, is_ab=False, alt=False,
                  rank_col=1, tier_col=3, score_col=2, iv_col=4,
                  action_col=None, url_col=None):
    """Writes one job row with full colour coding."""
    bg_hex, fg_hex = TIER_COLORS.get(tier, ("E0E0E0","000000"))
    row_bg = AB_GREEN if is_ab else (GRAY2 if alt else GRAY1)

    for ci, val in enumerate(vals, 1):
        if ci == rank_col:
            data_cell(ws, row, ci, val, "1E3A5F", WHITE, bold=True, center=True)
        elif ci == score_col:
            data_cell(ws, row, ci, val, bg_hex, fg_hex, bold=True, center=True)
        elif ci == tier_col:
            data_cell(ws, row, ci, val, bg_hex, fg_hex, bold=True, center=True)
        elif ci == iv_col:
            data_cell(ws, row, ci, val, bg_hex, fg_hex, center=True)
        elif action_col and ci == action_col:
            data_cell(ws, row, ci, val, bg_hex, fg_hex, bold=True, center=True)
        elif url_col and ci == url_col:
            data_cell(ws, row, ci, val, row_bg, "1155CC", underline=True)
        else:
            data_cell(ws, row, ci, val, row_bg)
    ws.row_dimensions[row].height = 50

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════
# RESUME PARSING
# ════════════════════════════════════════════════════════════════════════
SKILL_LIBRARY = {
    "salesforce":               {"keywords":["salesforce"],                          "weight":25,"category":"platform"},
    "sales cloud":              {"keywords":["sales cloud"],                         "weight":8, "category":"platform"},
    "service cloud":            {"keywords":["service cloud"],                       "weight":8, "category":"platform"},
    "financial services cloud": {"keywords":["financial services cloud"," fsc"],     "weight":12,"category":"platform"},
    "marketing cloud":          {"keywords":["marketing cloud"],                     "weight":12,"category":"platform"},
    "experience cloud":         {"keywords":["experience cloud","community cloud"],  "weight":7, "category":"platform"},
    "salesforce cpq":           {"keywords":["salesforce cpq"," cpq"],              "weight":8, "category":"platform"},
    "ncino":                    {"keywords":["ncino","n-cino"],                      "weight":12,"category":"platform"},
    "loan iq":                  {"keywords":["loan iq"],                             "weight":8, "category":"platform"},
    "hotdocs":                  {"keywords":["hotdocs"],                             "weight":6, "category":"platform"},
    "crm":                      {"keywords":["crm","customer relationship"],         "weight":5, "category":"platform"},
    "ms dynamics":              {"keywords":["dynamics 365","ms dynamics"],          "weight":7, "category":"platform"},
    "hubspot":                  {"keywords":["hubspot"],                             "weight":7, "category":"platform"},
    "functional testing":       {"keywords":["functional testing","functional test"],"weight":8, "category":"qa"},
    "regression testing":       {"keywords":["regression testing","regression test"],"weight":8, "category":"qa"},
    "integration testing":      {"keywords":["integration testing"],                 "weight":8, "category":"qa"},
    "uat":                      {"keywords":["uat","user acceptance"],               "weight":6, "category":"qa"},
    "system testing":           {"keywords":["system testing","system test"],        "weight":5, "category":"qa"},
    "smoke testing":            {"keywords":["smoke testing","smoke test","sanity test"],"weight":4,"category":"qa"},
    "api testing":              {"keywords":["api testing","postman","rest api"],     "weight":7, "category":"qa"},
    "performance testing":      {"keywords":["performance testing","load testing"],  "weight":6, "category":"qa"},
    "security testing":         {"keywords":["security testing","profiles","owd","fls"],"weight":6,"category":"qa"},
    "selenium":                 {"keywords":["selenium"],                            "weight":7, "category":"automation"},
    "accelq":                   {"keywords":["accelq"],                              "weight":8, "category":"automation"},
    "playwright":               {"keywords":["playwright"],                          "weight":7, "category":"automation"},
    "cypress":                  {"keywords":["cypress"],                             "weight":7, "category":"automation"},
    "appium":                   {"keywords":["appium"],                              "weight":6, "category":"automation"},
    "tosca":                    {"keywords":["tosca"],                               "weight":7, "category":"automation"},
    "katalon":                  {"keywords":["katalon"],                             "weight":6, "category":"automation"},
    "jira":                     {"keywords":["jira"],                               "weight":4, "category":"tool"},
    "hp alm":                   {"keywords":["hp alm","quality center"," alm"],     "weight":4, "category":"tool"},
    "azure devops":             {"keywords":["azure devops"," ado "],               "weight":5, "category":"tool"},
    "postman":                  {"keywords":["postman"],                             "weight":6, "category":"tool"},
    "copado":                   {"keywords":["copado"],                              "weight":5, "category":"tool"},
    "git":                      {"keywords":["bitbucket","github"," git "],         "weight":3, "category":"tool"},
    "soql":                     {"keywords":["soql"],                               "weight":7, "category":"technical"},
    "sql":                      {"keywords":[" sql ","sql,","sql\n"],               "weight":5, "category":"technical"},
    "apex":                     {"keywords":["apex"],                               "weight":6, "category":"technical"},
    "python":                   {"keywords":["python"],                             "weight":5, "category":"technical"},
    "agile":                    {"keywords":["agile","scrum","sprint"],             "weight":4, "category":"methodology"},
    "rtm":                      {"keywords":["traceability","rtm"],                 "weight":4, "category":"methodology"},
    "test planning":            {"keywords":["test plan","test strategy","test case"],"weight":5,"category":"methodology"},
    "defect management":        {"keywords":["defect","bug tracking"],              "weight":3, "category":"methodology"},
    "banking":                  {"keywords":["banking"," bank ","financial services","fintech"],"weight":10,"category":"domain"},
    "insurance":                {"keywords":["insurance"],                          "weight":8, "category":"domain"},
    "wealth management":        {"keywords":["wealth management","investment"],     "weight":8, "category":"domain"},
    "lending":                  {"keywords":["lending","loan","mortgage","commercial lending"],"weight":8,"category":"domain"},
    "healthcare":               {"keywords":["healthcare"," health ","hl7","fhir"], "weight":7, "category":"domain"},
    "retail":                   {"keywords":["retail","e-commerce","ecommerce"],    "weight":6, "category":"domain"},
    "telecom":                  {"keywords":["telecom","telecommunications"],       "weight":6, "category":"domain"},
    "team lead":                {"keywords":["team lead","led team","mentored","managed team"],"weight":5,"category":"seniority"},
    "qa lead":                  {"keywords":["qa lead","qe lead","test lead"],      "weight":5, "category":"seniority"},
    "senior":                   {"keywords":["senior","sr."],                       "weight":3, "category":"seniority"},
}

def extract_text_from_pdf(path):
    try:
        import pypdf
        text = ""
        with open(path, "rb") as f:
            for page in pypdf.PdfReader(f).pages:
                text += page.extract_text() or ""
        return text
    except Exception as e:
        print(f"  PDF read error: {e}"); return ""

def extract_text_from_docx(path):
    try:
        import docx
        return "\n".join(p.text for p in docx.Document(path).paragraphs)
    except Exception as e:
        print(f"  DOCX read error: {e}"); return ""

def load_resume():
    resume_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resumes")
    os.makedirs(resume_dir, exist_ok=True)
    found = []
    for pat in ["*.pdf","*.PDF","*.docx","*.DOCX","*.doc"]:
        found.extend(glob.glob(os.path.join(resume_dir, pat)))
    if not found:
        print(f"⚠️  No resume in {resume_dir}"); return None
    path = sorted(found, key=os.path.getmtime, reverse=True)[0]
    print(f"📄 Resume: {os.path.basename(path)}")
    ext = os.path.splitext(path)[1].lower()
    return extract_text_from_pdf(path) if ext == ".pdf" else extract_text_from_docx(path)

def parse_resume(text):
    if not text: return {}
    tl = text.lower()
    detected = {}
    for skill, data in SKILL_LIBRARY.items():
        for kw in data["keywords"]:
            if kw in tl:
                detected[skill] = data.copy(); break
    yoe = re.search(r'(\d+)\+?\s*years?\s*(of\s+)?experience', tl)
    name = text.strip().split('\n')[0].strip()
    name = name if len(name) < 50 else "Candidate"
    platforms = [s for s,d in detected.items() if d["category"]=="platform"]
    domains   = [s for s,d in detected.items() if d["category"]=="domain"]
    print(f"\n👤 {name} | {int(yoe.group(1)) if yoe else 0}+ yrs | {len(detected)} skills detected")
    print(f"   Platforms: {', '.join(platforms[:6])}")
    print(f"   Domain: {', '.join(domains[:4])}")
    return {"name":name, "years":int(yoe.group(1)) if yoe else 0,
            "skills":detected, "platform_skills":platforms, "domain_skills":domains,
            "raw_text":text}

def build_search_queries(profile):
    skills = profile.get("skills",{})
    platforms = profile.get("platform_skills",[])
    domains   = profile.get("domain_skills",[])
    queries = []
    platform_display = {"salesforce":"Salesforce","ms dynamics":"Dynamics 365",
                        "hubspot":"HubSpot","crm":"CRM"}
    for p in platforms[:2]:
        dp = platform_display.get(p, p.title())
        for qa in ["QA","Quality Assurance","QA Analyst","Quality Engineer"]:
            queries.append(f"{dp} {qa}")
    if "ncino" in skills: queries += ["nCino QA","nCino testing"]
    for d in domains[:2]:
        queries.append(f"QA analyst {d}")
    for tool in ["selenium","playwright","accelq","tosca"]:
        if tool in skills: queries.append(f"QA automation {tool}"); break
    if not queries: queries = ["QA Engineer","Quality Assurance Analyst"]
    seen = set(); unique = []
    for q in queries:
        if q.lower() not in seen: seen.add(q.lower()); unique.append(q)
    print(f"\n🔍 Queries: {', '.join(unique[:8])}")
    return unique[:8]


# ════════════════════════════════════════════════════════════════════════
# DYNAMIC SCORING
# ════════════════════════════════════════════════════════════════════════
def score_job(job, profile):
    title = job.get("title","").lower()
    desc  = str(job.get("descriptionText","")).lower()
    comb  = title + " " + desc
    score = 0; matches = []; gaps = []
    skills = profile.get("skills",{})

    for skill, data in skills.items():
        for kw in data["keywords"]:
            if kw in comb:
                score += data["weight"]; matches.append(skill.title()); break

    years = profile.get("years",0)
    if years >= 7 and any(w in title for w in ["senior","lead","principal","manager","director","head"]):
        score += 5; matches.append("Seniority match")
    elif years >= 3 and any(w in title for w in ["intermediate","mid","analyst"]):
        score += 3; matches.append("Level match")

    gap_checks = {"Playwright":["playwright"],"Mobile/Appium":["appium","ios testing","android testing"],
                  "SAP":["sap "],"Guidewire":["guidewire"],"French req":["bilingue","bilingual french"],
                  "Non-IT QA":["food safety","haccp","manufacturing quality","field qc","construction quality"]}
    for gname, gkws in gap_checks.items():
        in_job = any(kw in comb for kw in gkws)
        in_res = any(kw in profile.get("raw_text","").lower() for kw in gkws)
        if in_job and not in_res: gaps.append(gname)

    is_ab = any(w in job.get("location","").lower() for w in
                ["alberta","calgary","edmonton","sherwood park","red deer"," ab,",", ab"])

    if score>=65:   tier,bg,iv="🏆 EXCELLENT","1A7340","Very High (85-95%)"
    elif score>=45: tier,bg,iv="✅ STRONG","2D6A2D","High (70-85%)"
    elif score>=25: tier,bg,iv="🟡 GOOD","7F6000","Medium (50-70%)"
    elif score>=12: tier,bg,iv="🟠 STRETCH","8B3A0F","Low-Medium (30-50%)"
    else:           tier,bg,iv="❌ LOW MATCH","5A0000","Low (<30%)"

    return dict(score=min(score,100), tier=tier, tier_bg=bg, interview=iv,
                matches=", ".join(dict.fromkeys(matches)) or "General match",
                gaps=", ".join(gaps) or "None identified", is_alberta=is_ab)


# ════════════════════════════════════════════════════════════════════════
# SCRAPER
# ════════════════════════════════════════════════════════════════════════
def scrape_jobs(queries):
    client = ApifyClient(os.environ["APIFY_API_TOKEN"])
    all_jobs = {}

    li_urls = []
    for q in queries[:6]:
        kw = q.replace(" ","+")
        for loc in ["Canada","Alberta%2C+Canada","Calgary%2C+Alberta%2C+Canada"]:
            li_urls.append(f"https://www.linkedin.com/jobs/search/?keywords={kw}&location={loc}&f_TPR=r86400&position=1&pageNum=0")

    print(f"\n🔗 LinkedIn: {len(li_urls)} searches...")
    try:
        run = client.actor("curious_coder/linkedin-jobs-scraper").call(
            run_input={"urls":li_urls[:8],"count":100,"scrapeCompany":False})
        for item in client.dataset(run["defaultDatasetId"]).iterate_items():
            jid = str(item.get("id",""))
            if jid and jid not in all_jobs:
                item["source"] = "LinkedIn"; all_jobs[jid] = item
        print(f"   ✅ {len(all_jobs)} jobs")
    except Exception as e: print(f"   LinkedIn error: {e}")

    print(f"\n🔗 Indeed: {len(queries[:4])} queries × 2 locations...")
    for q in queries[:4]:
        for loc in ["Alberta","Canada"]:
            try:
                run = client.actor("borderline/indeed-scraper").call(
                    run_input={"country":"ca","query":q,"location":loc,
                               "maxRows":30,"sort":"date","enableUniqueJobs":True})
                for item in client.dataset(run["defaultDatasetId"]).iterate_items():
                    t=item.get("title",""); co=item.get("company","")
                    key=f"indeed_{t}_{co}_{loc}"
                    if key not in all_jobs and t:
                        lc=item.get("location",{}); lc=lc.get("formattedAddressShort","") if isinstance(lc,dict) else str(lc)
                        sal=item.get("salary","") or ""; sal=sal.get("salaryText","") if isinstance(sal,dict) else str(sal)
                        all_jobs[key]={"id":key,"title":t,"companyName":co,"location":str(lc),
                                       "postedAt":item.get("date",""),"salary":str(sal),
                                       "descriptionText":str(item.get("description","") or item.get("snippet","")),
                                       "link":item.get("url","") or item.get("externalApplyLink",""),
                                       "source":"Indeed"}
            except Exception as e: print(f"   Indeed error ({q}/{loc}): {e}")

    print(f"\n✅ Total scraped: {len(all_jobs)}")
    return list(all_jobs.values())


# ════════════════════════════════════════════════════════════════════════
# DEDUP
# ════════════════════════════════════════════════════════════════════════
def get_new_jobs(jobs):
    seen = set()
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f: seen = set(json.load(f))
    new=[]; new_keys=set()
    for j in jobs:
        k=f"{j.get('title','').lower().strip()}|{j.get('companyName','').lower().strip()}"
        if k not in seen: new.append(j); new_keys.add(k)
    seen.update(new_keys)
    with open(SEEN_FILE,"w") as f: json.dump(list(seen),f)
    print(f"   New (unseen): {len(new)}")
    return new


# ════════════════════════════════════════════════════════════════════════
# MULTI-SHEET EXCEL BUILDER
# ════════════════════════════════════════════════════════════════════════
def build_excel(scored_jobs, profile, timestamp):
    name       = profile.get("name","Candidate")
    first_name = name.split()[0]
    skill_summary = ", ".join(list(profile.get("skills",{}).keys())[:14])

    ab_jobs   = [j for j in scored_jobs if j.get("is_alberta")]
    top_jobs  = [j for j in scored_jobs if j["score"] >= 45]
    tc        = {}
    for j in scored_jobs: tc[j["tier"]] = tc.get(j["tier"],0)+1

    wb = Workbook()

    # ──────────────────────────────────────────────────────────────────
    # SHEET 1 — 🆕 All New Jobs
    # ──────────────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "🆕 All New Jobs"

    banner(ws1,1,15,
        f"🆕 JOB ALERT — {timestamp} | {name.upper()} | Resume-Powered | {len(scored_jobs)} New Matches",
        sz=11, height=28)
    banner(ws1,2,15,
        f"Skills from resume: {skill_summary} | 🌲 Green rows = Alberta jobs",
        bg="1C2E40", fg=WHITE, sz=9, height=16)

    hdrs1 = ["#","🌲AB","Score","Match Tier","Interview %","Job Title","Company",
             "Location","Source","Posted","Salary","Employment","Key Matches","Action","Apply Link"]
    col_headers(ws1, 3, hdrs1)

    for i, j in enumerate(scored_jobs, 1):
        ab  = j.get("is_alberta", False)
        ab_badge = "🌲 AB" if ab else "—"
        vals = [i, ab_badge, f"{j['score']}/100", j["tier"], j["interview"],
                j.get("title",""), j.get("companyName",""), j.get("location",""),
                j.get("source",""), j.get("postedAt",""),
                j.get("salary","") or "Not listed",
                j.get("employmentType","") or "N/A",
                j.get("matches",""), PRI.get(j["tier"],""),
                j.get("link","") or ""]
        write_job_row(ws1, 3+i, vals, j["tier"], ab, i%2==0,
                      rank_col=1, score_col=3, tier_col=4, iv_col=5,
                      action_col=14, url_col=15)
        # Override AB badge cell colour
        c = ws1.cell(row=3+i, column=2)
        c.fill = PatternFill("solid", start_color=AB_DARK if ab else GRAY2)
        c.font = Font(name="Arial",size=9,bold=ab,color=WHITE if ab else "888888")
        c.alignment = Alignment(horizontal="center",vertical="center")

    set_col_widths(ws1, [4,7,9,16,18,40,26,26,10,11,16,12,48,15,55])
    ws1.freeze_panes = "A4"

    # ──────────────────────────────────────────────────────────────────
    # SHEET 2 — 🌲 Alberta Jobs Only
    # ──────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🌲 Alberta Jobs Only")

    banner(ws2,1,13,
        f"🌲 ALBERTA JOBS ONLY — Calgary & Edmonton | {len(ab_jobs)} jobs | {timestamp}",
        bg="0A3D1F", sz=11, height=28)
    banner(ws2,2,13,
        "Apply today while applicant counts are still low! All rows matched against your resume.",
        bg="122E1A", fg="A8D5A2", sz=9, height=16)

    hdrs2 = ["#","Score","Match Tier","Interview %","Job Title","Company",
             "City","Source","Posted","Salary","Key Matches","Gaps","Apply Link"]
    col_headers(ws2, 3, hdrs2, bg="0A3D1F")

    for i, j in enumerate(ab_jobs, 1):
        city = j.get("location","").split(",")[0].strip()
        vals = [i, f"{j['score']}/100", j["tier"], j["interview"],
                j.get("title",""), j.get("companyName",""), city,
                j.get("source",""), j.get("postedAt",""),
                j.get("salary","") or "Not listed",
                j.get("matches",""), j.get("gaps","None"),
                j.get("link","") or ""]
        write_job_row(ws2, 3+i, vals, j["tier"], True, i%2==0,
                      rank_col=1, score_col=2, tier_col=3, iv_col=4, url_col=13)

    set_col_widths(ws2, [4,9,16,18,40,28,20,10,11,16,48,30,55])
    ws2.freeze_panes = "A4"

    # ──────────────────────────────────────────────────────────────────
    # SHEET 3 — 🏆 Top Matches (score ≥ 45)
    # ──────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("🏆 Top Matches – Apply Now")

    banner(ws3,1,11,
        f"🏆 TOP {len(top_jobs)} MATCHES (Score 45+) — Highest Interview Probability | {timestamp}")
    banner(ws3,2,11,
        f"🌲 = Alberta job | All matches scored dynamically from {first_name}'s resume",
        bg="1C2E40", fg=WHITE, sz=9, height=16)

    hdrs3 = ["#","🌲","Score","Tier","Job Title","Company","Location",
             "Source","Interview %","Key Matches","Apply Link"]
    col_headers(ws3, 3, hdrs3)

    for i, j in enumerate(top_jobs, 1):
        ab  = j.get("is_alberta", False)
        vals = [i, "🌲" if ab else "—",
                f"{j['score']}/100", j["tier"], j.get("title",""),
                j.get("companyName",""), j.get("location",""),
                j.get("source",""), j["interview"],
                j.get("matches",""), j.get("link","") or ""]
        write_job_row(ws3, 3+i, vals, j["tier"], ab, i%2==0,
                      rank_col=1, score_col=3, tier_col=4, iv_col=9, url_col=11)
        c = ws3.cell(row=3+i, column=2)
        c.fill = PatternFill("solid", start_color=AB_DARK if ab else GRAY2)
        c.font = Font(name="Arial",size=9,bold=ab,color=WHITE if ab else "888888")
        c.alignment = Alignment(horizontal="center",vertical="center")

    set_col_widths(ws3, [4,5,9,16,40,28,28,10,18,52,55])
    ws3.freeze_panes = "A4"

    # ──────────────────────────────────────────────────────────────────
    # SHEET 4 — 📊 Dashboard
    # ──────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("📊 Dashboard")
    banner(ws4,1,5, f"📊 JOB ALERT DASHBOARD — {name.upper()} | {timestamp}", sz=13, height=34)

    # Stats block
    cal = sum(1 for j in scored_jobs if "calgary" in j.get("location","").lower())
    edm = sum(1 for j in scored_jobs if "edmonton" in j.get("location","").lower()
              or "sherwood" in j.get("location","").lower())
    stats = [
        ("📋 New Jobs Found This Alert",         len(scored_jobs)),
        ("🏆 Excellent Matches (65+)",            tc.get("🏆 EXCELLENT",0)),
        ("✅ Strong Matches (45–64)",             tc.get("✅ STRONG",0)),
        ("🟡 Good Matches (25–44)",               tc.get("🟡 GOOD",0)),
        ("🟠 Stretch Roles (12–24)",              tc.get("🟠 STRETCH",0)),
        ("❌ Low/No Match (<12)",                 tc.get("❌ LOW MATCH",0)),
        ("🌲 Alberta Total (Calgary + Edmonton)", len(ab_jobs)),
        ("📍 Calgary",                            cal),
        ("📍 Edmonton / Sherwood Park",           edm),
        ("🍁 Canada-Wide (non-Alberta)",          len(scored_jobs)-len(ab_jobs)),
        ("🔗 From LinkedIn",                      sum(1 for j in scored_jobs if j.get("source")=="LinkedIn")),
        ("🔍 From Indeed",                        sum(1 for j in scored_jobs if j.get("source")=="Indeed")),
        ("⏰ Alert Generated",                    timestamp),
    ]
    stat_bgs = [DARK_NAVY,"1A7340","2D6A2D","7F6000","8B3A0F","5A0000",
                "0A3D1F","0A4D2A","0A4D2A","1E3A5F","2E4A6F","2E4A6F","444444"]

    for ci,v in [(1,"METRIC"),(2,"COUNT")]:
        c = ws4.cell(row=2, column=ci, value=v)
        c.fill = PatternFill("solid",start_color="1E3A5F")
        c.font = Font(bold=True,color=GOLD,name="Arial",size=10)
        c.alignment = Alignment(horizontal="center"); tb(c)
    ws4.row_dimensions[2].height = 24

    for i,(label,count) in enumerate(stats,3):
        for ci,val in [(1,label),(2,str(count))]:
            c = ws4.cell(row=i,column=ci,value=val)
            c.fill = PatternFill("solid",start_color=stat_bgs[i-3])
            c.font = Font(name="Arial",size=10,color=WHITE,bold=(i==3))
            c.alignment = Alignment(vertical="center",horizontal="center" if ci==2 else "left",wrap_text=True)
            tb(c)
        ws4.row_dimensions[i].height = 22

    # Top 12 shortlist
    banner(ws4, 17, 5, "🏆 TOP 12 JOBS TO APPLY TODAY", sz=11, height=26)
    col_headers(ws4, 18, ["Score","Tier","Job Title","Company","City / Location"], bg="1E3A5F", height=22)

    for i, j in enumerate(scored_jobs[:12], 1):
        bg_hex = TIER_COLORS.get(j["tier"],("E0E0E0",""))[0]
        ab     = j.get("is_alberta",False)
        city   = j.get("location","").split(",")[0].strip()
        prefix = "🌲 " if ab else ""
        for ci, val in enumerate([f"{j['score']}/100", j["tier"],
                                   prefix+j.get("title",""),
                                   j.get("companyName",""), city], 1):
            c = ws4.cell(row=18+i,column=ci,value=str(val))
            c.fill = PatternFill("solid",start_color=bg_hex)
            c.font = Font(name="Arial",size=9,color=WHITE,bold=(ci<=2))
            c.alignment = Alignment(vertical="center",wrap_text=True,
                                    horizontal="center" if ci<=2 else "left")
            tb(c)
        ws4.row_dimensions[18+i].height = 26

    # Resume skills box
    banner(ws4, 32, 5, "📄 RESUME SKILLS DETECTED — Used for Scoring", sz=11, height=26)
    skill_lines = [f"{'⭐'*min(d['weight']//4+1,5)} {s.title()} ({d['category'].title()}, {d['weight']}pts)"
                   for s,d in sorted(profile.get("skills",{}).items(), key=lambda x:-x[1]["weight"])]
    for i,line in enumerate(skill_lines[:20],33):
        ws4.merge_cells(f"A{i}:E{i}")
        c = ws4.cell(row=i,column=1,value=line)
        c.fill = PatternFill("solid",start_color="1C2E40" if i%2==0 else "162438")
        c.font = Font(name="Arial",size=9,color=WHITE)
        c.alignment = Alignment(vertical="center",wrap_text=True)
        tb(c); ws4.row_dimensions[i].height = 20

    # Tips
    tips_start = 54
    banner(ws4, tips_start, 5, "💡 APPLY STRATEGY TIPS", sz=11, height=26)
    tips = [
        f"1. 🌲 Alberta-first: Calgary & Edmonton have {len(ab_jobs)} matches — apply there first for local advantage",
        "2. Your rarest combo is SF FSC + nCino + Banking domain — lead every application with this in your summary",
        "3. Apply within 48 hrs of posting — early applicants get 3–5× more recruiter attention on LinkedIn",
        f"4. Top match score this run: {scored_jobs[0]['score']}/100 — {scored_jobs[0].get('title','')} at {scored_jobs[0].get('companyName','')}",
        "5. Update resume? Just replace the file in /resumes folder — next run will re-detect all skills automatically",
        "6. Set LinkedIn alerts: 'Salesforce QA' + 'Canada' + 'Daily' to catch postings between your 6-hour runs",
    ]
    for i,tip in enumerate(tips, tips_start+1):
        ws4.merge_cells(f"A{i}:E{i}")
        c = ws4.cell(row=i,column=1,value=tip)
        c.fill = PatternFill("solid",start_color="1C2E40" if i%2==0 else "162438")
        c.font = Font(name="Arial",size=9,color=WHITE)
        c.alignment = Alignment(vertical="center",wrap_text=True)
        tb(c); ws4.row_dimensions[i].height = 28

    set_col_widths(ws4, [52,12,40,28,24])

    fname = f"JobAlert_{first_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    print(f"✅ Excel built: {fname}")
    return fname


# ════════════════════════════════════════════════════════════════════════
# EMAIL
# ════════════════════════════════════════════════════════════════════════
def send_email(filepath, scored_jobs, profile):
    name   = profile.get("name","there")
    gmail  = os.environ["GMAIL_ADDRESS"]
    pwd    = os.environ["GMAIL_APP_PASSWORD"]
    to     = os.environ.get("RECIPIENT_EMAIL", gmail)
    ab_cnt = sum(1 for j in scored_jobs if j.get("is_alberta"))
    top_cnt= sum(1 for j in scored_jobs if j["score"]>=45)

    msg = MIMEMultipart()
    msg["From"]    = gmail; msg["To"] = to
    msg["Subject"] = (f"🆕 {len(scored_jobs)} New Job Matches for {name.split()[0]} "
                      f"— {datetime.now().strftime('%b %d, %I:%M %p')} "
                      f"| {ab_cnt} Alberta 🌲 | {top_cnt} Top Matches 🏆")

    # Top 5 rows for email body
    rows_html = ""
    for j in scored_jobs[:5]:
        ab_flag = "🌲 " if j.get("is_alberta") else ""
        tier_color = {"🏆 EXCELLENT":"#1A7340","✅ STRONG":"#2D6A2D",
                      "🟡 GOOD":"#7F6000","🟠 STRETCH":"#8B3A0F"}.get(j["tier"],"#555")
        rows_html += f"""
        <tr>
          <td style="padding:8px;border:1px solid #ddd;text-align:center">
            <span style="background:{tier_color};color:white;padding:3px 8px;border-radius:10px;font-size:12px;font-weight:bold">{j['score']}/100</span>
          </td>
          <td style="padding:8px;border:1px solid #ddd"><b>{ab_flag}{j.get('title','')}</b></td>
          <td style="padding:8px;border:1px solid #ddd">{j.get('companyName','')}</td>
          <td style="padding:8px;border:1px solid #ddd">{j.get('location','')}</td>
          <td style="padding:8px;border:1px solid #ddd;font-size:11px;color:#555">{j.get('matches','')[:55]}</td>
          <td style="padding:8px;border:1px solid #ddd;text-align:center">
            <a href="{j.get('link','')}" style="background:#1E3A5F;color:white;padding:5px 10px;border-radius:5px;text-decoration:none;font-size:12px">Apply →</a>
          </td>
        </tr>"""

    skill_badges = "".join(
        f'<span style="background:#1E3A5F;color:white;padding:2px 8px;border-radius:10px;'
        f'margin:2px;display:inline-block;font-size:11px">{s.title()}</span>'
        for s in list(profile.get("skills",{}).keys())[:16]
    )
    summary_row = lambda label, val, col: (
        f'<td style="padding:8px;text-align:center;background:{col};color:white;border-radius:5px;margin:4px">'
        f'<div style="font-size:22px;font-weight:bold">{val}</div>'
        f'<div style="font-size:11px">{label}</div></td>'
    )

    html = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:920px;margin:0 auto;background:#f0f2f5">
      <!-- Header -->
      <div style="background:#0D1B2A;padding:24px 28px;border-radius:10px 10px 0 0">
        <h2 style="color:#C9A84C;margin:0;font-size:20px">🆕 Job Alert — {name.split()[0]}</h2>
        <p style="color:#aaa;margin:6px 0 0;font-size:13px">{datetime.now().strftime('%A, %B %d %Y at %I:%M %p')} · Resume-powered matching</p>
      </div>
      <!-- Skill badges -->
      <div style="background:#1C2E40;padding:12px 28px">
        <p style="color:#C9A84C;margin:0 0 6px;font-size:12px">📄 Skills detected from your resume:</p>
        <div>{skill_badges}</div>
      </div>
      <!-- Summary stats -->
      <div style="background:#fff;padding:16px 28px;border:1px solid #ddd">
        <table style="width:100%;border-collapse:separate;border-spacing:6px">
          <tr>
            {summary_row("New Jobs Found", len(scored_jobs), "#0D1B2A")}
            {summary_row("🏆 Top Matches", top_cnt, "#1A7340")}
            {summary_row("🌲 Alberta Jobs", ab_cnt, "#0A3D1F")}
            {summary_row("Excellent (65+)", sum(1 for j in scored_jobs if j['score']>=65), "#2D6A2D")}
          </tr>
        </table>
      </div>
      <!-- Top 5 table -->
      <div style="background:#fff;padding:20px 28px;border:1px solid #ddd;border-top:none">
        <h3 style="color:#0D1B2A;margin:0 0 12px">🏆 Top 5 Matches This Run</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <tr style="background:#1E3A5F;color:#C9A84C">
            <th style="padding:10px;border:1px solid #ddd">Score</th>
            <th style="padding:10px;border:1px solid #ddd">Job Title</th>
            <th style="padding:10px;border:1px solid #ddd">Company</th>
            <th style="padding:10px;border:1px solid #ddd">Location</th>
            <th style="padding:10px;border:1px solid #ddd">Matched Skills</th>
            <th style="padding:10px;border:1px solid #ddd">Apply</th>
          </tr>
          {rows_html}
        </table>
      </div>
      <!-- Footer -->
      <div style="background:#1C2E40;padding:14px 28px;border-radius:0 0 10px 10px">
        <p style="color:#aaa;margin:0;font-size:12px">
          📎 <b style="color:#C9A84C">Excel attached</b> with 4 sheets:
          🆕 All New Jobs · 🌲 Alberta Only · 🏆 Top Matches · 📊 Dashboard<br>
          🌲 Green rows in Excel = Alberta (Calgary/Edmonton) jobs<br>
          ⏰ Next alert in 6 hours · Update resume: replace file in <code>/resumes</code> folder on GitHub
        </p>
      </div>
    </body></html>"""

    msg.attach(MIMEText(html, "html"))
    with open(filepath,"rb") as f:
        part = MIMEBase("application","octet-stream"); part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",f"attachment; filename={os.path.basename(filepath)}")
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
        s.login(gmail,pwd); s.sendmail(gmail,to,msg.as_string())
    print(f"✅ Email sent → {to}")


# ════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    ts = datetime.now().strftime("%b %d %Y, %I:%M %p")
    print(f"\n{'='*65}\n  🔍 RESUME-POWERED JOB ALERT v2.0 — {ts}\n{'='*65}")

    resume_text = load_resume()
    if not resume_text:
        print("❌ No resume found. Add PDF/DOCX to /resumes folder."); sys.exit(1)

    profile  = parse_resume(resume_text)
    queries  = build_search_queries(profile)
    raw_jobs = scrape_jobs(queries)
    new_jobs = get_new_jobs(raw_jobs)

    scored = []
    for j in new_jobs:
        result = score_job(j, profile)
        j.update(result)
        if result["score"] >= 12:
            scored.append(j)
    scored.sort(key=lambda x: -x["score"])

    print(f"\n📊 {len(scored)} matching new jobs")
    if not scored:
        print("ℹ️  No new matches this run — skipping email."); sys.exit(0)

    xl = build_excel(scored, profile, ts)
    send_email(xl, scored, profile)
    os.remove(xl)
    print(f"\n✅ Done! Alert sent for {len(scored)} jobs.")
