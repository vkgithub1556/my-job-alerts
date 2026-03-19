"""
RESUME-POWERED JOB ALERT SYSTEM v2.0
- Drop your resume (PDF or DOCX) into the /resumes folder
- Script auto-reads it, extracts skills, builds search queries
- Scrapes LinkedIn + Indeed + Google Jobs every 6 hours via GitHub Actions
- Emails you a 4-sheet Excel with ONLY NEW jobs each run
- Never sends the same job twice (seen_jobs.json memory)

SHEETS IN EXCEL:
  Sheet 1 - All New Jobs (every new job ranked by match score)
  Sheet 2 - Alberta Jobs Only (Calgary / Edmonton)
  Sheet 3 - Top Matches (score 45+, highest interview probability)
  Sheet 4 - Dashboard (stats, top 12 shortlist, resume skills, tips)
"""

import os
import json
import re
import smtplib
import sys
import glob
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from apify_client import ApifyClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONSTANTS
# ============================================================
DARK_NAVY = "0D1B2A"
GOLD      = "C9A84C"
WHITE     = "FFFFFF"
GRAY1     = "F5F5F5"
GRAY2     = "EAEAEA"
AB_GREEN  = "E8F5E9"
AB_DARK   = "1A5F3A"

TIER_COLORS = {
    "EXCELLENT": ("1A7340", "FFFFFF"),
    "STRONG":    ("2D6A2D", "FFFFFF"),
    "GOOD":      ("7F6000", "FFFFFF"),
    "STRETCH":   ("8B3A0F", "FFFFFF"),
    "LOW":       ("5A0000", "FFFFFF"),
}
TIER_LABELS = {
    "EXCELLENT": "EXCELLENT",
    "STRONG":    "STRONG",
    "GOOD":      "GOOD",
    "STRETCH":   "STRETCH",
    "LOW":       "LOW MATCH",
}
PRI_LABELS = {
    "EXCELLENT": "APPLY NOW",
    "STRONG":    "Apply - Tailor CV",
    "GOOD":      "Worth Applying",
    "STRETCH":   "Optional",
    "LOW":       "Skip",
}

SEEN_FILE = "seen_jobs.json"

# ============================================================
# SKILL LIBRARY - system checks which ones appear in resume
# ============================================================
SKILL_LIBRARY = {
    "salesforce":               {"keywords": ["salesforce"],                           "weight": 25, "category": "platform"},
    "sales cloud":              {"keywords": ["sales cloud"],                          "weight": 8,  "category": "platform"},
    "service cloud":            {"keywords": ["service cloud"],                        "weight": 8,  "category": "platform"},
    "financial services cloud": {"keywords": ["financial services cloud", " fsc"],     "weight": 12, "category": "platform"},
    "marketing cloud":          {"keywords": ["marketing cloud"],                      "weight": 12, "category": "platform"},
    "experience cloud":         {"keywords": ["experience cloud", "community cloud"],  "weight": 7,  "category": "platform"},
    "salesforce cpq":           {"keywords": ["salesforce cpq", " cpq "],             "weight": 8,  "category": "platform"},
    "ncino":                    {"keywords": ["ncino", "n-cino"],                      "weight": 12, "category": "platform"},
    "loan iq":                  {"keywords": ["loan iq"],                              "weight": 8,  "category": "platform"},
    "hotdocs":                  {"keywords": ["hotdocs"],                              "weight": 6,  "category": "platform"},
    "crm":                      {"keywords": ["crm", "customer relationship"],         "weight": 5,  "category": "platform"},
    "ms dynamics":              {"keywords": ["dynamics 365", "ms dynamics"],          "weight": 7,  "category": "platform"},
    "hubspot":                  {"keywords": ["hubspot"],                              "weight": 7,  "category": "platform"},
    "functional testing":       {"keywords": ["functional testing", "functional test"],"weight": 8,  "category": "qa"},
    "regression testing":       {"keywords": ["regression testing", "regression test"],"weight": 8,  "category": "qa"},
    "integration testing":      {"keywords": ["integration testing"],                  "weight": 8,  "category": "qa"},
    "uat":                      {"keywords": ["uat", "user acceptance"],               "weight": 6,  "category": "qa"},
    "system testing":           {"keywords": ["system testing", "system test"],        "weight": 5,  "category": "qa"},
    "smoke testing":            {"keywords": ["smoke testing", "smoke test", "sanity test"], "weight": 4, "category": "qa"},
    "api testing":              {"keywords": ["api testing", "postman", "rest api"],   "weight": 7,  "category": "qa"},
    "performance testing":      {"keywords": ["performance testing", "load testing"],  "weight": 6,  "category": "qa"},
    "security testing":         {"keywords": ["security testing", "profiles", "owd", "fls"], "weight": 6, "category": "qa"},
    "selenium":                 {"keywords": ["selenium"],                             "weight": 7,  "category": "automation"},
    "accelq":                   {"keywords": ["accelq"],                               "weight": 8,  "category": "automation"},
    "playwright":               {"keywords": ["playwright"],                           "weight": 7,  "category": "automation"},
    "cypress":                  {"keywords": ["cypress"],                              "weight": 7,  "category": "automation"},
    "appium":                   {"keywords": ["appium"],                               "weight": 6,  "category": "automation"},
    "tosca":                    {"keywords": ["tosca"],                                "weight": 7,  "category": "automation"},
    "katalon":                  {"keywords": ["katalon"],                              "weight": 6,  "category": "automation"},
    "jira":                     {"keywords": ["jira"],                                 "weight": 4,  "category": "tool"},
    "hp alm":                   {"keywords": ["hp alm", "quality center", " alm "],   "weight": 4,  "category": "tool"},
    "azure devops":             {"keywords": ["azure devops", " ado "],               "weight": 5,  "category": "tool"},
    "postman":                  {"keywords": ["postman"],                              "weight": 6,  "category": "tool"},
    "copado":                   {"keywords": ["copado"],                               "weight": 5,  "category": "tool"},
    "git":                      {"keywords": ["bitbucket", "github", " git "],        "weight": 3,  "category": "tool"},
    "soql":                     {"keywords": ["soql"],                                 "weight": 7,  "category": "technical"},
    "sql":                      {"keywords": [" sql ", "sql,", "sql\n"],              "weight": 5,  "category": "technical"},
    "apex":                     {"keywords": ["apex"],                                 "weight": 6,  "category": "technical"},
    "python":                   {"keywords": ["python"],                               "weight": 5,  "category": "technical"},
    "agile":                    {"keywords": ["agile", "scrum", "sprint"],             "weight": 4,  "category": "methodology"},
    "rtm":                      {"keywords": ["traceability", "rtm"],                  "weight": 4,  "category": "methodology"},
    "test planning":            {"keywords": ["test plan", "test strategy", "test case"], "weight": 5, "category": "methodology"},
    "defect management":        {"keywords": ["defect", "bug tracking"],               "weight": 3,  "category": "methodology"},
    "banking":                  {"keywords": ["banking", " bank ", "financial services", "fintech"], "weight": 10, "category": "domain"},
    "insurance":                {"keywords": ["insurance"],                            "weight": 8,  "category": "domain"},
    "wealth management":        {"keywords": ["wealth management", "investment"],      "weight": 8,  "category": "domain"},
    "lending":                  {"keywords": ["lending", "loan", "mortgage", "commercial lending"], "weight": 8, "category": "domain"},
    "healthcare":               {"keywords": ["healthcare", " health ", "hl7", "fhir"],"weight": 7,  "category": "domain"},
    "retail":                   {"keywords": ["retail", "e-commerce", "ecommerce"],   "weight": 6,  "category": "domain"},
    "telecom":                  {"keywords": ["telecom", "telecommunications"],        "weight": 6,  "category": "domain"},
    "team lead":                {"keywords": ["team lead", "led team", "mentored", "managed team"], "weight": 5, "category": "seniority"},
    "qa lead":                  {"keywords": ["qa lead", "qe lead", "test lead"],     "weight": 5,  "category": "seniority"},
    "senior":                   {"keywords": ["senior", "sr."],                        "weight": 3,  "category": "seniority"},
}

# ============================================================
# EXCEL HELPER FUNCTIONS
# ============================================================
def add_border(cell):
    side = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def make_banner(ws, row, total_cols, text, bg=DARK_NAVY, fg=GOLD, font_size=12, row_height=28):
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(bold=True, color=fg, name="Arial", size=font_size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = row_height


def make_header_row(ws, row, headers, bg="1E3A5F", fg=GOLD, row_height=34):
    ws.row_dimensions[row].height = row_height
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.font = Font(bold=True, color=fg, name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        add_border(cell)


def write_job_row(ws, row_num, values, tier_key, is_alberta=False, alternate=False,
                  rank_col=1, score_col=2, tier_col=3, iv_col=4,
                  action_col=None, url_col=None):
    bg_hex, fg_hex = TIER_COLORS.get(tier_key, ("E0E0E0", "000000"))
    if is_alberta:
        row_bg = AB_GREEN
    elif alternate:
        row_bg = GRAY2
    else:
        row_bg = GRAY1

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=str(value) if value is not None else "")
        add_border(cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if col_idx == rank_col:
            cell.fill = PatternFill("solid", start_color="1E3A5F")
            cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col_idx in (score_col, tier_col):
            cell.fill = PatternFill("solid", start_color=bg_hex)
            cell.font = Font(name="Arial", size=9, bold=True, color=fg_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col_idx == iv_col:
            cell.fill = PatternFill("solid", start_color=bg_hex)
            cell.font = Font(name="Arial", size=9, color=fg_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif action_col and col_idx == action_col:
            cell.fill = PatternFill("solid", start_color=bg_hex)
            cell.font = Font(name="Arial", size=9, bold=True, color=fg_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        elif url_col and col_idx == url_col:
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")

        else:
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.font = Font(name="Arial", size=9, color="000000")

    ws.row_dimensions[row_num].height = 50


def set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ============================================================
# RESUME LOADING
# ============================================================
def extract_pdf_text(filepath):
    try:
        import pypdf
        text = ""
        with open(filepath, "rb") as f:
            reader = pypdf.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() or ""
        return text
    except Exception as e:
        print(f"  PDF error: {e}")
        return ""


def extract_docx_text(filepath):
    try:
        import docx
        doc = docx.Document(filepath)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        print(f"  DOCX error: {e}")
        return ""


def load_resume():
    resume_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resumes")
    os.makedirs(resume_dir, exist_ok=True)

    found_files = []
    for pattern in ["*.pdf", "*.PDF", "*.docx", "*.DOCX", "*.doc"]:
        found_files.extend(glob.glob(os.path.join(resume_dir, pattern)))

    if not found_files:
        print(f"No resume found in {resume_dir}")
        print("Please add a PDF or DOCX file to the /resumes folder")
        return None

    # Use most recently modified resume
    resume_path = sorted(found_files, key=os.path.getmtime, reverse=True)[0]
    print(f"Resume: {os.path.basename(resume_path)}")

    extension = os.path.splitext(resume_path)[1].lower()
    if extension == ".pdf":
        return extract_pdf_text(resume_path)
    elif extension in (".docx", ".doc"):
        return extract_docx_text(resume_path)
    return ""


# ============================================================
# RESUME PARSING - extract skills profile
# ============================================================
def parse_resume(resume_text):
    if not resume_text:
        return {}

    text_lower = resume_text.lower()
    detected_skills = {}

    for skill_name, skill_data in SKILL_LIBRARY.items():
        for keyword in skill_data["keywords"]:
            if keyword in text_lower:
                detected_skills[skill_name] = skill_data.copy()
                break

    # Extract years of experience
    yoe_match = re.search(r"(\d+)\+?\s*years?\s*(of\s+)?experience", text_lower)
    years_exp = int(yoe_match.group(1)) if yoe_match else 0

    # Extract candidate name from first line
    first_line = resume_text.strip().split("\n")[0].strip()
    candidate_name = first_line if len(first_line) < 50 else "Candidate"

    platform_skills = [s for s, d in detected_skills.items() if d["category"] == "platform"]
    domain_skills   = [s for s, d in detected_skills.items() if d["category"] == "domain"]
    qa_skills       = [s for s, d in detected_skills.items() if d["category"] == "qa"]

    print(f"Name: {candidate_name}")
    print(f"Experience: {years_exp}+ years")
    print(f"Platform skills: {', '.join(platform_skills[:6])}")
    print(f"Domain: {', '.join(domain_skills[:4])}")
    print(f"QA skills: {', '.join(qa_skills[:5])}")
    print(f"Total skills detected: {len(detected_skills)}")

    return {
        "name":            candidate_name,
        "years":           years_exp,
        "skills":          detected_skills,
        "platform_skills": platform_skills,
        "domain_skills":   domain_skills,
        "qa_skills":       qa_skills,
        "raw_text":        resume_text,
    }


# ============================================================
# BUILD SEARCH QUERIES FROM RESUME
# ============================================================
def build_search_queries(profile):
    detected = profile.get("skills", {})
    platforms = profile.get("platform_skills", [])
    domains   = profile.get("domain_skills", [])

    queries = []

    platform_display_names = {
        "salesforce": "Salesforce",
        "ms dynamics": "Dynamics 365",
        "hubspot": "HubSpot",
        "crm": "CRM",
    }

    qa_terms = ["QA", "Quality Assurance", "QA Analyst", "Quality Engineer"]

    for platform in platforms[:2]:
        display_name = platform_display_names.get(platform, platform.title())
        for qa_term in qa_terms[:3]:
            queries.append(f"{display_name} {qa_term}")

    if "ncino" in detected:
        queries.append("nCino QA")
        queries.append("nCino testing")

    for domain in domains[:2]:
        queries.append(f"QA analyst {domain}")

    for automation_tool in ["selenium", "playwright", "accelq", "tosca"]:
        if automation_tool in detected:
            queries.append(f"QA automation {automation_tool}")
            break

    if not queries:
        queries = ["QA Engineer", "Quality Assurance Analyst", "Software Tester"]

    # Deduplicate
    seen_lower = set()
    unique_queries = []
    for q in queries:
        if q.lower() not in seen_lower:
            seen_lower.add(q.lower())
            unique_queries.append(q)

    print(f"Search queries: {', '.join(unique_queries[:8])}")
    return unique_queries[:8]


# ============================================================
# JOB SCORING - dynamic based on resume
# ============================================================
def score_job(job, profile):
    title    = job.get("title", "").lower()
    desc     = str(job.get("descriptionText", "")).lower()
    combined = title + " " + desc

    score = 0
    matches = []
    gaps = []
    detected_skills = profile.get("skills", {})

    # Score each detected resume skill against job description
    for skill_name, skill_data in detected_skills.items():
        for keyword in skill_data["keywords"]:
            if keyword in combined:
                score += skill_data["weight"]
                matches.append(skill_name.title())
                break

    # Seniority bonus
    years = profile.get("years", 0)
    if years >= 7:
        if any(w in title for w in ["senior", "lead", "principal", "manager", "director", "head"]):
            score += 5
            matches.append("Seniority match")
    elif years >= 3:
        if any(w in title for w in ["intermediate", "mid", "analyst"]):
            score += 3
            matches.append("Level match")

    # Gap detection - skills in job but NOT in resume
    gap_checks = {
        "Playwright":    ["playwright"],
        "Mobile/Appium": ["appium", "ios testing", "android testing"],
        "SAP":           ["sap "],
        "Guidewire":     ["guidewire"],
        "French req":    ["bilingue", "bilingual french"],
        "Non-IT QA":     ["food safety", "haccp", "manufacturing quality", "field qc"],
    }
    for gap_name, gap_keywords in gap_checks.items():
        in_job    = any(kw in combined for kw in gap_keywords)
        in_resume = any(kw in profile.get("raw_text", "").lower() for kw in gap_keywords)
        if in_job and not in_resume:
            gaps.append(gap_name)

    # Alberta flag
    location_lower = job.get("location", "").lower()
    is_alberta = any(w in location_lower for w in [
        "alberta", "calgary", "edmonton", "sherwood park", "red deer", " ab,", ", ab"
    ])

    # Determine tier
    if score >= 65:
        tier_key = "EXCELLENT"
        interview = "Very High (85-95%)"
    elif score >= 45:
        tier_key = "STRONG"
        interview = "High (70-85%)"
    elif score >= 25:
        tier_key = "GOOD"
        interview = "Medium (50-70%)"
    elif score >= 12:
        tier_key = "STRETCH"
        interview = "Low-Medium (30-50%)"
    else:
        tier_key = "LOW"
        interview = "Low (<30%)"

    return {
        "score":      min(score, 100),
        "tier_key":   tier_key,
        "tier_label": TIER_LABELS[tier_key],
        "interview":  interview,
        "matches":    ", ".join(dict.fromkeys(matches)) or "General match",
        "gaps":       ", ".join(gaps) or "None identified",
        "is_alberta": is_alberta,
    }


# ============================================================
# SCRAPER - LinkedIn + Indeed via Apify
# ============================================================
def scrape_jobs(search_queries):
    """
    Scrapes jobs from 3 sources:
      1. LinkedIn  - professional network, Canada + Alberta, past 24 hours
      2. Indeed    - broad job board, Canada + Alberta, sorted by date
      3. Google Jobs - meta-aggregator pulling from LinkedIn, Indeed, Glassdoor,
                       company career pages, Workopolis, Monster, and 100s more
    All results are merged and deduplicated by title+company key.
    """
    client = ApifyClient(os.environ["APIFY_API_TOKEN"])
    all_jobs = {}

    # ----------------------------------------------------------
    # SOURCE 1: LinkedIn
    # ----------------------------------------------------------
    linkedin_urls = []
    for query in search_queries[:6]:
        encoded_query = query.replace(" ", "+")
        for location in ["Canada", "Alberta%2C+Canada", "Calgary%2C+Alberta%2C+Canada"]:
            url = (f"https://www.linkedin.com/jobs/search/"
                   f"?keywords={encoded_query}&location={location}"
                   f"&f_TPR=r86400&position=1&pageNum=0")
            linkedin_urls.append(url)

    print(f"Source 1 - LinkedIn: {len(linkedin_urls)} searches")
    try:
        run = client.actor("curious_coder/linkedin-jobs-scraper").call(
            run_input={"urls": linkedin_urls[:8], "count": 100, "scrapeCompany": False}
        )
        for item in client.dataset(run["defaultDatasetId"]).iterate_items():
            job_id = str(item.get("id", ""))
            if job_id and job_id not in all_jobs:
                item["source"] = "LinkedIn"
                all_jobs[job_id] = item
        print(f"  LinkedIn: {len(all_jobs)} jobs found")
    except Exception as e:
        print(f"  LinkedIn error: {e}")

    count_after_linkedin = len(all_jobs)

    # ----------------------------------------------------------
    # SOURCE 2: Indeed Canada
    # ----------------------------------------------------------
    print(f"Source 2 - Indeed: {len(search_queries[:4])} queries x 2 locations")
    for query in search_queries[:4]:
        for location in ["Alberta", "Canada"]:
            try:
                run = client.actor("borderline/indeed-scraper").call(
                    run_input={
                        "country": "ca",
                        "query": query,
                        "location": location,
                        "maxRows": 30,
                        "sort": "date",
                        "enableUniqueJobs": True,
                    }
                )
                for item in client.dataset(run["defaultDatasetId"]).iterate_items():
                    title   = item.get("title", "")
                    company = item.get("company", "")
                    key     = f"indeed_{title}_{company}_{location}"
                    if key not in all_jobs and title:
                        loc = item.get("location", {})
                        if isinstance(loc, dict):
                            loc = loc.get("formattedAddressShort", "") or "Canada"
                        sal = item.get("salary", "") or ""
                        if isinstance(sal, dict):
                            sal = sal.get("salaryText", "") or ""
                        all_jobs[key] = {
                            "id":              key,
                            "title":           title,
                            "companyName":     company,
                            "location":        str(loc),
                            "postedAt":        item.get("date", ""),
                            "salary":          str(sal),
                            "descriptionText": str(item.get("description", "") or item.get("snippet", "")),
                            "link":            item.get("url", "") or item.get("externalApplyLink", ""),
                            "source":          "Indeed",
                            "employmentType":  str(item.get("jobType", "") or ""),
                        }
            except Exception as e:
                print(f"  Indeed error ({query}/{location}): {e}")

    count_after_indeed = len(all_jobs)
    print(f"  Indeed: {count_after_indeed - count_after_linkedin} new jobs added")

    # ----------------------------------------------------------
    # SOURCE 3: Google Jobs
    # Meta-aggregator: pulls from LinkedIn, Indeed, Glassdoor,
    # company career pages, Workopolis, Monster, Eluta, and more.
    # Input schema: query (str), location (str), country (str),
    #               language (str), google_domain (str),
    #               num_results (int), max_pagination (int)
    # Output: list of dicts under key "jobs" inside each result item,
    #         each job has: title, company_name, location, description,
    #         salary, job_type, posted_date, application_link, requirements
    # ----------------------------------------------------------
    google_locations = [
        ("Canada", "ca"),
        ("Calgary, Alberta", "ca"),
        ("Edmonton, Alberta", "ca"),
    ]
    print(f"Source 3 - Google Jobs: {len(search_queries[:5])} queries x {len(google_locations)} locations")
    google_new = 0
    for query in search_queries[:5]:
        for location, country_code in google_locations:
            try:
                run = client.actor("johnvc/google-jobs-scraper").call(
                    run_input={
                        "query":         query,
                        "location":      location,
                        "country":       country_code,
                        "language":      "en",
                        "google_domain": "google.ca",
                        "num_results":   50,
                        "max_pagination": 3,
                    }
                )
                for result_item in client.dataset(run["defaultDatasetId"]).iterate_items():
                    # Google Jobs actor returns a wrapper object with a "jobs" list inside
                    jobs_list = result_item.get("jobs", [])
                    # Also handle flat format (some versions return jobs directly)
                    if not jobs_list and result_item.get("title"):
                        jobs_list = [result_item]

                    for job in jobs_list:
                        title   = job.get("title", "")
                        company = job.get("company_name", "") or job.get("company", "")
                        if not title:
                            continue
                        key = f"google_{title.lower().strip()}_{company.lower().strip()}"
                        if key not in all_jobs:
                            # Normalize posted_date - Google uses relative strings like "2 days ago"
                            posted = job.get("posted_date", "") or ""
                            # Convert salary - may be a string or dict
                            sal = job.get("salary", "") or ""
                            if isinstance(sal, dict):
                                sal = sal.get("salaryText", "") or ""
                            # Requirements list -> join to string for description enrichment
                            reqs = job.get("requirements", [])
                            reqs_text = " Requirements: " + " | ".join(reqs) if reqs else ""
                            all_jobs[key] = {
                                "id":              key,
                                "title":           title,
                                "companyName":     company,
                                "location":        job.get("location", location),
                                "postedAt":        posted,
                                "salary":          str(sal),
                                "descriptionText": str(job.get("description", "")) + reqs_text,
                                "link":            job.get("application_link", "") or "",
                                "source":          "Google Jobs",
                                "employmentType":  str(job.get("job_type", "") or ""),
                            }
                            google_new += 1
            except Exception as e:
                print(f"  Google Jobs error ({query}/{location}): {e}")

    print(f"  Google Jobs: {google_new} new jobs added")
    print(f"Total unique jobs scraped (all 3 sources): {len(all_jobs)}")
    return list(all_jobs.values())


# ============================================================
# DEDUPLICATION - never send same job twice
# ============================================================
def get_new_jobs(jobs):
    seen_keys = set()
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            seen_keys = set(json.load(f))

    new_jobs  = []
    new_keys  = set()
    for job in jobs:
        key = f"{job.get('title', '').lower().strip()}|{job.get('companyName', '').lower().strip()}"
        if key not in seen_keys:
            new_jobs.append(job)
            new_keys.add(key)

    # Save updated memory
    seen_keys.update(new_keys)
    with open(SEEN_FILE, "w") as f:
        json.dump(list(seen_keys), f)

    print(f"New (never seen before): {len(new_jobs)}")
    return new_jobs


# ============================================================
# EXCEL BUILDER - 4 sheets
# ============================================================
def build_excel(scored_jobs, profile, timestamp):
    candidate_name = profile.get("name", "Candidate")
    first_name     = candidate_name.split()[0]
    skill_summary  = ", ".join(list(profile.get("skills", {}).keys())[:14])

    alberta_jobs = [j for j in scored_jobs if j.get("is_alberta")]
    top_jobs     = [j for j in scored_jobs if j["score"] >= 45]

    tier_counts = {}
    for job in scored_jobs:
        tier_counts[job["tier_key"]] = tier_counts.get(job["tier_key"], 0) + 1

    wb = Workbook()

    # ===========================================================
    # SHEET 1 - All New Jobs
    # ===========================================================
    ws1 = wb.active
    ws1.title = "All New Jobs"

    make_banner(ws1, 1, 15,
        f"JOB ALERT - {timestamp} | {candidate_name.upper()} | Resume-Powered | {len(scored_jobs)} New Matches",
        font_size=11, row_height=28)
    make_banner(ws1, 2, 15,
        f"Skills from resume: {skill_summary} | Green rows = Alberta jobs",
        bg="1C2E40", fg=WHITE, font_size=9, row_height=16)

    headers1 = [
        "#", "AB", "Score", "Match Tier", "Interview %", "Job Title", "Company",
        "Location", "Source", "Posted", "Salary", "Employment", "Key Matches", "Action", "Apply Link"
    ]
    make_header_row(ws1, 3, headers1)

    for i, job in enumerate(scored_jobs, 1):
        is_ab   = job.get("is_alberta", False)
        ab_flag = "AB" if is_ab else "-"
        values  = [
            i, ab_flag, f"{job['score']}/100", job["tier_label"], job["interview"],
            job.get("title", ""), job.get("companyName", ""), job.get("location", ""),
            job.get("source", ""), job.get("postedAt", ""),
            job.get("salary", "") or "Not listed",
            job.get("employmentType", "") or "N/A",
            job.get("matches", ""), PRI_LABELS.get(job["tier_key"], ""),
            job.get("link", "") or "",
        ]
        write_job_row(ws1, 3 + i, values, job["tier_key"], is_ab, i % 2 == 0,
                      rank_col=1, score_col=3, tier_col=4, iv_col=5,
                      action_col=14, url_col=15)
        # Color the AB badge cell
        ab_cell = ws1.cell(row=3 + i, column=2)
        ab_cell.fill = PatternFill("solid", start_color=AB_DARK if is_ab else GRAY2)
        ab_cell.font = Font(name="Arial", size=9, bold=is_ab, color=WHITE if is_ab else "888888")
        ab_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_column_widths(ws1, [4, 6, 9, 16, 18, 40, 26, 26, 10, 11, 16, 12, 48, 15, 55])
    ws1.freeze_panes = "A4"

    # ===========================================================
    # SHEET 2 - Alberta Jobs Only
    # ===========================================================
    ws2 = wb.create_sheet("Alberta Jobs Only")

    make_banner(ws2, 1, 13,
        f"ALBERTA JOBS ONLY - Calgary & Edmonton | {len(alberta_jobs)} Jobs | {timestamp}",
        bg="0A3D1F", font_size=11, row_height=28)
    make_banner(ws2, 2, 13,
        "All rows matched against your resume. Apply early while applicant counts are low!",
        bg="122E1A", fg="A8D5A2", font_size=9, row_height=16)

    headers2 = [
        "#", "Score", "Match Tier", "Interview %", "Job Title", "Company",
        "City", "Source", "Posted", "Salary", "Key Matches", "Gaps", "Apply Link"
    ]
    make_header_row(ws2, 3, headers2, bg="0A3D1F")

    for i, job in enumerate(alberta_jobs, 1):
        city   = job.get("location", "").split(",")[0].strip()
        values = [
            i, f"{job['score']}/100", job["tier_label"], job["interview"],
            job.get("title", ""), job.get("companyName", ""), city,
            job.get("source", ""), job.get("postedAt", ""),
            job.get("salary", "") or "Not listed",
            job.get("matches", ""), job.get("gaps", "None"),
            job.get("link", "") or "",
        ]
        write_job_row(ws2, 3 + i, values, job["tier_key"], True, i % 2 == 0,
                      rank_col=1, score_col=2, tier_col=3, iv_col=4, url_col=13)

    set_column_widths(ws2, [4, 9, 16, 18, 40, 28, 20, 10, 11, 16, 48, 30, 55])
    ws2.freeze_panes = "A4"

    # ===========================================================
    # SHEET 3 - Top Matches (score 45+)
    # ===========================================================
    ws3 = wb.create_sheet("Top Matches - Apply Now")

    make_banner(ws3, 1, 11,
        f"TOP {len(top_jobs)} MATCHES (Score 45+) - Highest Interview Probability | {timestamp}",
        font_size=11, row_height=28)
    make_banner(ws3, 2, 11,
        f"AB = Alberta job | All scores calculated dynamically from {first_name}'s resume",
        bg="1C2E40", fg=WHITE, font_size=9, row_height=16)

    headers3 = [
        "#", "AB", "Score", "Tier", "Job Title", "Company",
        "Location", "Source", "Interview %", "Key Matches", "Apply Link"
    ]
    make_header_row(ws3, 3, headers3)

    for i, job in enumerate(top_jobs, 1):
        is_ab  = job.get("is_alberta", False)
        values = [
            i, "AB" if is_ab else "-",
            f"{job['score']}/100", job["tier_label"], job.get("title", ""),
            job.get("companyName", ""), job.get("location", ""),
            job.get("source", ""), job["interview"],
            job.get("matches", ""), job.get("link", "") or "",
        ]
        write_job_row(ws3, 3 + i, values, job["tier_key"], is_ab, i % 2 == 0,
                      rank_col=1, score_col=3, tier_col=4, iv_col=9, url_col=11)
        ab_cell = ws3.cell(row=3 + i, column=2)
        ab_cell.fill = PatternFill("solid", start_color=AB_DARK if is_ab else GRAY2)
        ab_cell.font = Font(name="Arial", size=9, bold=is_ab, color=WHITE if is_ab else "888888")
        ab_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_column_widths(ws3, [4, 5, 9, 16, 40, 28, 28, 10, 18, 52, 55])
    ws3.freeze_panes = "A4"

    # ===========================================================
    # SHEET 4 - Dashboard
    # ===========================================================
    ws4 = wb.create_sheet("Dashboard")

    make_banner(ws4, 1, 5,
        f"JOB ALERT DASHBOARD - {candidate_name.upper()} | {timestamp}",
        font_size=13, row_height=34)

    # Stats
    calgary_count  = sum(1 for j in scored_jobs if "calgary" in j.get("location", "").lower())
    edmonton_count = sum(1 for j in scored_jobs if "edmonton" in j.get("location", "").lower()
                         or "sherwood" in j.get("location", "").lower())

    stats_rows = [
        ("New Jobs Found This Alert",            len(scored_jobs)),
        ("Excellent Matches (65+)",              tier_counts.get("EXCELLENT", 0)),
        ("Strong Matches (45-64)",               tier_counts.get("STRONG", 0)),
        ("Good Matches (25-44)",                 tier_counts.get("GOOD", 0)),
        ("Stretch Roles (12-24)",                tier_counts.get("STRETCH", 0)),
        ("Low or No Match (below 12)",           tier_counts.get("LOW", 0)),
        ("Alberta Total (Calgary + Edmonton)",   len(alberta_jobs)),
        ("Calgary",                              calgary_count),
        ("Edmonton / Sherwood Park",             edmonton_count),
        ("Canada-Wide (non-Alberta)",            len(scored_jobs) - len(alberta_jobs)),
        ("From LinkedIn",                        sum(1 for j in scored_jobs if j.get("source") == "LinkedIn")),
        ("From Indeed",                          sum(1 for j in scored_jobs if j.get("source") == "Indeed")),
        ("From Google Jobs",                     sum(1 for j in scored_jobs if j.get("source") == "Google Jobs")),
        ("Alert Generated",                      timestamp),
    ]
    stat_bg_colors = [
        DARK_NAVY, "1A7340", "2D6A2D", "7F6000", "8B3A0F", "5A0000",
        "0A3D1F", "0A4D2A", "0A4D2A", "1E3A5F", "2E4A6F", "2E4A6F", "1A3A6F", "444444"
    ]

    # Stats header
    for col_idx, text in [(1, "METRIC"), (2, "COUNT")]:
        cell = ws4.cell(row=2, column=col_idx, value=text)
        cell.fill = PatternFill("solid", start_color="1E3A5F")
        cell.font = Font(bold=True, color=GOLD, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center")
        add_border(cell)
    ws4.row_dimensions[2].height = 24

    for row_idx, (label, count) in enumerate(stats_rows, 3):
        row_bg = stat_bg_colors[row_idx - 3]
        for col_idx, value in [(1, label), (2, str(count))]:
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.font = Font(name="Arial", size=10, color=WHITE, bold=(row_idx == 3))
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col_idx == 2 else "left",
                wrap_text=True,
            )
            add_border(cell)
        ws4.row_dimensions[row_idx].height = 22

    # Top 12 shortlist
    make_banner(ws4, 17, 5, "TOP 12 JOBS TO APPLY TODAY", font_size=11, row_height=26)
    make_header_row(ws4, 18, ["Score", "Tier", "Job Title", "Company", "City / Location"],
                    row_height=22)

    for i, job in enumerate(scored_jobs[:12], 1):
        bg_hex = TIER_COLORS.get(job["tier_key"], ("E0E0E0", ""))[0]
        is_ab  = job.get("is_alberta", False)
        city   = job.get("location", "").split(",")[0].strip()
        prefix = "AB - " if is_ab else ""
        for col_idx, value in enumerate([
            f"{job['score']}/100", job["tier_label"],
            prefix + job.get("title", ""),
            job.get("companyName", ""), city
        ], 1):
            cell = ws4.cell(row=18 + i, column=col_idx, value=str(value))
            cell.fill = PatternFill("solid", start_color=bg_hex)
            cell.font = Font(name="Arial", size=9, color=WHITE, bold=(col_idx <= 2))
            cell.alignment = Alignment(
                vertical="center", wrap_text=True,
                horizontal="center" if col_idx <= 2 else "left",
            )
            add_border(cell)
        ws4.row_dimensions[18 + i].height = 26

    # Resume skills block
    make_banner(ws4, 32, 5, "RESUME SKILLS DETECTED - Used for Scoring", font_size=11, row_height=26)
    skills_sorted = sorted(profile.get("skills", {}).items(), key=lambda x: -x[1]["weight"])
    skill_bg_colors = {
        "platform": "1A7340", "qa": "2D6A2D", "automation": "7F6000",
        "tool": "1E3A5F", "technical": "2E4A6F", "methodology": "3A5A3A",
        "domain": "5A1A00", "seniority": "0D1B2A",
    }
    for row_idx, (skill_name, skill_data) in enumerate(skills_sorted[:20], 33):
        skill_bg = skill_bg_colors.get(skill_data["category"], "444444")
        stars    = "*" * min(skill_data["weight"] // 4 + 1, 5)
        line     = f"{stars} {skill_name.title()} ({skill_data['category'].title()}, {skill_data['weight']} pts)"
        ws4.merge_cells(f"A{row_idx}:E{row_idx}")
        cell = ws4.cell(row=row_idx, column=1, value=line)
        cell.fill = PatternFill("solid", start_color=skill_bg)
        cell.font = Font(name="Arial", size=9, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        add_border(cell)
        ws4.row_dimensions[row_idx].height = 20

    # Tips block
    tips_start = 55
    make_banner(ws4, tips_start, 5, "APPLY STRATEGY TIPS", font_size=11, row_height=26)
    tips = [
        f"1. Alberta-first strategy: Calgary & Edmonton have {len(alberta_jobs)} matches - apply there first for local advantage",
        "2. Your rarest combo is Salesforce FSC + nCino + Banking domain - lead every application with this",
        "3. Apply within 48 hrs of posting - early applicants get 3-5x more recruiter attention on LinkedIn",
        f"4. Top match this alert: {scored_jobs[0]['score']}/100 - {scored_jobs[0].get('title','')} at {scored_jobs[0].get('companyName','')}",
        "5. Update resume? Replace the file in /resumes folder on GitHub - next run re-detects all skills automatically",
        "6. LinkedIn tip: Set job alerts for 'Salesforce QA' + Canada + Daily to catch postings between 6-hour runs",
    ]
    for row_idx, tip in enumerate(tips, tips_start + 1):
        ws4.merge_cells(f"A{row_idx}:E{row_idx}")
        cell = ws4.cell(row=row_idx, column=1, value=tip)
        cell.fill = PatternFill("solid", start_color="1C2E40" if row_idx % 2 == 0 else "162438")
        cell.font = Font(name="Arial", size=9, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        add_border(cell)
        ws4.row_dimensions[row_idx].height = 28

    set_column_widths(ws4, [52, 12, 40, 28, 24])

    # Save file
    filename = f"JobAlert_{first_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"Excel saved: {filename}")
    return filename


# ============================================================
# EMAIL SENDER
# ============================================================
def send_email(filepath, scored_jobs, profile):
    candidate_name = profile.get("name", "there")
    first_name     = candidate_name.split()[0]
    gmail_address  = os.environ["GMAIL_ADDRESS"]
    gmail_password = os.environ["GMAIL_APP_PASSWORD"]
    recipient      = os.environ.get("RECIPIENT_EMAIL", gmail_address)

    alberta_count = sum(1 for j in scored_jobs if j.get("is_alberta"))
    top_count     = sum(1 for j in scored_jobs if j["score"] >= 45)
    excellent_count = sum(1 for j in scored_jobs if j["score"] >= 65)

    msg = MIMEMultipart()
    msg["From"]    = gmail_address
    msg["To"]      = recipient
    msg["Subject"] = (
        f"[Job Alert] {len(scored_jobs)} New Matches for {first_name} - "
        f"{datetime.now().strftime('%b %d %I:%M %p')} | "
        f"{alberta_count} Alberta | {top_count} Top Matches"
    )

    # Build top 5 rows for email table
    top5_html = ""
    for job in scored_jobs[:5]:
        tier_bg_map = {
            "EXCELLENT": "#1A7340", "STRONG": "#2D6A2D",
            "GOOD": "#7F6000", "STRETCH": "#8B3A0F", "LOW": "#5A0000",
        }
        score_bg  = tier_bg_map.get(job["tier_key"], "#555555")
        ab_prefix = "AB - " if job.get("is_alberta") else ""
        top5_html += f"""
        <tr>
          <td style="padding:8px;border:1px solid #ddd;text-align:center">
            <span style="background:{score_bg};color:white;padding:3px 10px;
                  border-radius:10px;font-size:12px;font-weight:bold">
              {job['score']}/100
            </span>
          </td>
          <td style="padding:8px;border:1px solid #ddd">
            <strong>{ab_prefix}{job.get('title','')}</strong>
          </td>
          <td style="padding:8px;border:1px solid #ddd">{job.get('companyName','')}</td>
          <td style="padding:8px;border:1px solid #ddd">{job.get('location','')}</td>
          <td style="padding:8px;border:1px solid #ddd;font-size:11px;color:#555">
            {job.get('matches','')[:60]}
          </td>
          <td style="padding:8px;border:1px solid #ddd;text-align:center">
            <a href="{job.get('link','')}"
               style="background:#1E3A5F;color:white;padding:5px 12px;
                      border-radius:5px;text-decoration:none;font-size:12px">
              Apply
            </a>
          </td>
        </tr>"""

    # Skill badges
    skill_badges_html = "".join(
        f'<span style="background:#1E3A5F;color:white;padding:2px 8px;'
        f'border-radius:10px;margin:2px;display:inline-block;font-size:11px">'
        f'{s.title()}</span>'
        for s in list(profile.get("skills", {}).keys())[:16]
    )

    # Summary stat boxes
    def stat_box(label, value, color):
        return (
            f'<td style="padding:10px;text-align:center;background:{color};'
            f'color:white;border-radius:6px;margin:4px">'
            f'<div style="font-size:22px;font-weight:bold">{value}</div>'
            f'<div style="font-size:11px;margin-top:3px">{label}</div></td>'
        )

    html_body = f"""
    <html>
    <body style="font-family:Arial,sans-serif;max-width:920px;margin:0 auto;background:#f0f2f5">

      <!-- Header -->
      <div style="background:#0D1B2A;padding:24px 28px;border-radius:10px 10px 0 0">
        <h2 style="color:#C9A84C;margin:0;font-size:20px">
          Job Alert for {first_name}
        </h2>
        <p style="color:#aaa;margin:6px 0 0;font-size:13px">
          {datetime.now().strftime('%A, %B %d %Y at %I:%M %p')} &nbsp;|&nbsp; Resume-powered matching
        </p>
      </div>

      <!-- Skill badges -->
      <div style="background:#1C2E40;padding:12px 28px">
        <p style="color:#C9A84C;margin:0 0 6px;font-size:12px">
          Skills detected from your resume:
        </p>
        <div>{skill_badges_html}</div>
      </div>

      <!-- Summary stats -->
      <div style="background:#fff;padding:16px 28px;border:1px solid #ddd">
        <table style="width:100%;border-collapse:separate;border-spacing:6px">
          <tr>
            {stat_box("New Jobs Found", len(scored_jobs), "#0D1B2A")}
            {stat_box("Top Matches", top_count, "#1A7340")}
            {stat_box("Alberta Jobs", alberta_count, "#0A3D1F")}
            {stat_box("Excellent (65+)", excellent_count, "#2D6A2D")}
          </tr>
        </table>
      </div>

      <!-- Top 5 table -->
      <div style="background:#fff;padding:20px 28px;border:1px solid #ddd;border-top:none">
        <h3 style="color:#0D1B2A;margin:0 0 12px">Top 5 Matches This Run</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <tr style="background:#1E3A5F;color:#C9A84C">
            <th style="padding:10px;border:1px solid #ddd">Score</th>
            <th style="padding:10px;border:1px solid #ddd">Job Title</th>
            <th style="padding:10px;border:1px solid #ddd">Company</th>
            <th style="padding:10px;border:1px solid #ddd">Location</th>
            <th style="padding:10px;border:1px solid #ddd">Matched Skills</th>
            <th style="padding:10px;border:1px solid #ddd">Apply</th>
          </tr>
          {top5_html}
        </table>
      </div>

      <!-- Footer -->
      <div style="background:#1C2E40;padding:14px 28px;border-radius:0 0 10px 10px">
        <p style="color:#aaa;margin:0;font-size:12px">
          Excel attached with 4 sheets:
          All New Jobs | Alberta Only | Top Matches | Dashboard<br>
          Green rows in Excel = Alberta (Calgary / Edmonton) jobs<br>
          Sources: LinkedIn, Indeed Canada, Google Jobs (aggregates 100+ boards) | Next alert in 6 hours.
          To update matching, replace your resume file in the /resumes folder on GitHub.
        </p>
      </div>

    </body>
    </html>"""

    msg.attach(MIMEText(html_body, "html"))

    # Attach Excel file
    with open(filepath, "rb") as f:
        attachment = MIMEBase("application", "octet-stream")
        attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(filepath)}"
        )
        msg.attach(attachment)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_address, gmail_password)
        server.sendmail(gmail_address, recipient, msg.as_string())

    print(f"Email sent to {recipient}")


# ============================================================
# MAIN ENTRY POINT
# ============================================================
if __name__ == "__main__":
    timestamp = datetime.now().strftime("%b %d %Y, %I:%M %p")
    print("=" * 60)
    print(f"  RESUME-POWERED JOB ALERT v2.0 - {timestamp}")
    print("=" * 60)

    # Step 1: Load resume
    resume_text = load_resume()
    if not resume_text:
        print("No resume found. Add a PDF or DOCX to the /resumes folder.")
        sys.exit(1)

    # Step 2: Parse skills from resume
    profile = parse_resume(resume_text)

    # Step 3: Build search queries from resume profile
    queries = build_search_queries(profile)

    # Step 4: Scrape LinkedIn + Indeed + Google Jobs
    raw_jobs = scrape_jobs(queries)

    # Step 5: Filter to only new (never sent before) jobs
    new_jobs = get_new_jobs(raw_jobs)

    # Step 6: Score each job against resume
    scored_jobs = []
    for job in new_jobs:
        result = score_job(job, profile)
        job.update(result)
        if result["score"] >= 12:
            scored_jobs.append(job)
    scored_jobs.sort(key=lambda x: -x["score"])

    print(f"Matching new jobs: {len(scored_jobs)}")

    if not scored_jobs:
        print("No new matching jobs this run - skipping email.")
        sys.exit(0)

    # Step 7: Build 4-sheet Excel
    excel_file = build_excel(scored_jobs, profile, timestamp)

    # Step 8: Send email with Excel attached
    send_email(excel_file, scored_jobs, profile)

    # Cleanup local Excel file
    os.remove(excel_file)

    print(f"Done! Alert sent for {len(scored_jobs)} new jobs.")
